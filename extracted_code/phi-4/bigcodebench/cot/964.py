import os
from pathlib import Path
import pandas as pd
import docx
from openpyxl import load_workbook

def task_func(source_directory: str, target_directory: str) -> int:
    if not os.path.exists(source_directory):
        raise FileNotFoundError(f"Source directory {source_directory} does not exist.")
    
    Path(target_directory).mkdir(parents=True, exist_ok=True)
    
    extensions = ('.txt', '.docx', '.xlsx', '.csv')
    files_to_convert = [f for f in os.listdir(source_directory) if f.endswith(extensions)]
    
    successful_conversions = 0
    
    for file_name in files_to_convert:
        source_path = os.path.join(source_directory, file_name)
        target_path = os.path.join(target_directory, os.path.splitext(file_name)[0] + '.csv')
        
        if file_name.endswith('.txt'):
            with open(source_path, 'r', encoding='utf-8') as file:
                text = file.read()
        elif file_name.endswith('.docx'):
            doc = docx.Document(source_path)
            text = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
        elif file_name.endswith('.xlsx'):
            workbook = load_workbook(filename=source_path, read_only=True)
            sheet = workbook.active
            text = '\n'.join(['\t'.join([str(cell.value) for cell in row]) for row in sheet.iter_rows()])
        elif file_name.endswith('.csv'):
            df = pd.read_csv(source_path)
            text = df.to_csv(index=False)
        
        df = pd.DataFrame([text], columns=['Text'])
        df.to_csv(target_path, index=False)
        successful_conversions += 1
    
    return successful_conversions