import re
import glob
import os
from openpyxl import load_workbook

def task_func(directory_path='./xlsx_files/'):
    double_quote_pattern = re.compile(r'"')
    excel_files = glob.glob(os.path.join(directory_path, '*.xlsx'))
    processed_files_count = 0

    for file_path in excel_files:
        workbook = load_workbook(file_path)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell.value = double_quote_pattern.sub(r'\"', str(cell.value))
        workbook.save(file_path)
        processed_files_count += 1

    return processed_files_count