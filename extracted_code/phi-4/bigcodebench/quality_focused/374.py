import regex as re
import glob
import os
from openpyxl import load_workbook

def task_func(directory_path='./xlsx_files/'):
    pattern = re.compile(r'"')
    count = 0

    for file_path in glob.glob(os.path.join(directory_path, '*.xlsx')):
        workbook = load_workbook(file_path)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = pattern.sub(r'\"', cell.value)
        workbook.save(file_path)
        count += 1

    return count