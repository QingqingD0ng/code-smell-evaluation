import regex as re
import glob
import os
from openpyxl import load_workbook

def task_func(directory_path='./xlsx_files/'):
    file_pattern = os.path.join(directory_path, '*.xlsx')
    files = glob.glob(file_pattern)
    count = 0

    for file_path in files:
        wb = load_workbook(filename=file_path)
        for sheet in wb:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = re.sub(r'"', r'\"', cell.value)
        wb.save(file_path)
        count += 1

    return count