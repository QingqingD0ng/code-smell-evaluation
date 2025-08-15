import os
from pathlib import Path
import pandas as pd
import docx
from openpyxl import load_workbook

def task_func(source_directory: str, target_directory: str) -> int:
    if not os.path.exists(source_directory):
        raise FileNotFoundError(f"The source directory {source_directory} does not exist.")
    
    if not os.path.exists(target_directory):
        os.makedirs(target_directory)
    
    extensions = ['.txt', '.docx', '.xlsx', '.csv']
    success_count = 0

    for file_path in Path(source_directory).glob('*'):
        if file_path.suffix in extensions:
            file_name = file_path.stem + '.csv'
            target_file_path = Path(target_directory) / file_name
            text_content = []

            if file_path.suffix == '.txt':
                with open(file_path, 'r', encoding='utf-8') as file:
                    text_content = file.readlines()

            elif file_path.suffix == '.docx':
                doc = docx.Document(file_path)
                text_content = [p.text for p in doc.paragraphs]

            elif file_path.suffix == '.xlsx':
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    text_content.append(row)

            elif file_path.suffix == '.csv':
                df = pd.read_csv(file_path)
                text_content = df.values.tolist()

            df_output = pd.DataFrame(text_content, columns=['Text'])
            df_output.to_csv(target_file_path, index=False)
            success_count += 1

    return success_count