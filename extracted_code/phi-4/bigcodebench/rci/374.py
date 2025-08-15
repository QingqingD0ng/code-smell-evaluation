import regex as re
import glob
import os
from openpyxl import load_workbook
import tempfile

DOUBLE_QUOTE = '"'

def task_func(directory_path='./xlsx_files/'):
    file_pattern = os.path.join(directory_path, '*.xlsx')
    files = glob.glob(file_pattern)
    processed_files_count = 0

    for file_path in files:
        try:
            wb = load_workbook(filename=file_path)
            for sheet in wb:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell.value = re.sub(DOUBLE_QUOTE, f'\\{DOUBLE_QUOTE}', cell.value)
            
            with tempfile.NamedTemporaryFile(delete=False, dir=directory_path) as temp_file:
                wb.save(temp_file.name)
                os.replace(temp_file.name, file_path)
            
            processed_files_count += 1
        except Exception as e:
            print(f"Error processing {file_path}: {e}")

    return processed_files_count