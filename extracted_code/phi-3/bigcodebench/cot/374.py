import regex as re

import glob

import os

from openpyxl import load_workbook


def task_func(directory_path='./xlsx_files/'):
    
    processed_files_count = 0
    
    for file_path in glob.glob(os.path.join(directory_path, '*.xlsx')):
    
        workbook = load_workbook(filename=file_path)
    
        for sheet in workbook.worksheets:
    
            for cell in sheet.iter_rows():
    
                for cell_value in cell:
    
                    matches = re.findall(r'"', str(cell_value.value))
    
                    if matches:
    
                        cell.value = re.sub(r'"', r'\"', str(cell_value.value))
    
        workbook.save(file_path)
    
        processed_files_count += 1
    
    return processed_files_count


# Example usage:

# result = task_func()