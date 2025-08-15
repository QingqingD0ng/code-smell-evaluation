import re

import glob

import os

from openpyxl import load_workbook


def task_func(directory_path='./xlsx_files/'):
    file_count = 0
    for file_path in glob.glob(os.path.join(directory_path, '*.xlsx')):
        file_count += 1
        workbook = load_workbook(file_path)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(min_row=1, max_col=1, max_row=worksheet.max_row, values_only=False):
                for cell in row:
                    if cell.value and '"' in cell.value:
                        cell.value = re.sub(r'"', '\\"', cell.value)
            workbook.save(file_path)
    return file_count


# Example usage:

# import tempfile
# temp_dir = tempfile.mkdtemp()
# workbook = Workbook()
# sheet = workbook.active
# sheet.append(['This is a "test" string.'])
# workbook.save(temp_dir + '/test.xlsx')
# print(task_func(temp_dir))