import os
import pandas as pd
from pathlib import Path
from docx import Document
from openpyxl import load_workbook

def task_func(source_directory: str, target_directory: str) -> int:
    source_path = Path(source_directory)
    target_path = Path(target_directory)

    if not source_path.exists():
        raise FileNotFoundError(f"Source directory '{source_directory}' does not exist.")

    target_path.mkdir(parents=True, exist_ok=True)

    extensions = ['.txt', '.docx', '.xlsx', '.csv']
    count = 0

    for file_path in source_path.rglob('*'):
        if file_path.suffix.lower() in extensions:
            target_file_path = target_path / f"{file_path.stem}.csv"
            if file_path.suffix.lower() in ['.txt', '.docx', '.xlsx']:
                if file_path.suffix.lower() == '.docx':
                    doc = Document(file_path)
                    text = "\n".join([para.text for para in doc.paragraphs])
                elif file_path.suffix.lower() == '.xlsx':
                    workbook = load_workbook(file_path)
                    sheet = workbook.active
                    text = sheet.cell(row=1, column=1).value
                else:
                    with open(file_path, 'r') as file:
                        text = file.read()
            else:
                text = file_path.read_text()

            df = pd.DataFrame([[text]], columns=['Text'])
            df.to_csv(target_file_path, index=False)
            count += 1

    return count

# Example usage:
# task_func('/Users/test/Documents', '/Users/test/Documents/csv_files')
# task_func('/path/to/source', '/path/to/target')