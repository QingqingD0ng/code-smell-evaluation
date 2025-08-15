import os
from pathlib import Path
import pandas as pd
from docx import Document
from openpyxl import load_workbook

def convert_to_csv(file_path, target_dir):
    _, file_extension = os.path.splitext(file_path)
    if file_extension.lower() in ['.txt', '.docx', '.xlsx', '.csv']:
        df = load_file_content_as_df(file_path)
        if df is not None:
            csv_path = target_dir / (Path(file_path).name.replace(file_extension, '.csv'))
            df.to_csv(csv_path, index=False)
            return 1
    return 0

def load_file_content_as_df(file_path):
    try:
        if file_path.lower().endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
            return pd.DataFrame({'Text': [text]})
        elif file_path.lower().endswith('.docx'):
            doc = Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
            return pd.DataFrame({'Text': [text]})
        elif file_path.lower().endswith('.xlsx'):
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            text = sheet.cell(row=1, column=1).value
            return pd.DataFrame({'Text': [text]})
        elif file_path.lower().endswith('.csv'):
            return pd.read_csv(file_path)
    except Exception as e:
        print(f"Error: {e}")
        return None

def task_func(source_directory: str, target_directory: str) -> int:
    source_path = Path(source_directory).resolve(strict=True)
    target_path = Path(target_directory).resolve(strict=True)
    if not source_path.exists() or not target_path.exists():
        raise FileNotFoundError(