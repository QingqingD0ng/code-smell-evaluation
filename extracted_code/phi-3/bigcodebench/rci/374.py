import re
import glob
import os
import tempfile
from openpyxl import Workbook, load_workbook
from copy import copy


def escape_quotes_in_string(input_string):
    return re.sub(r'"', r'\"', input_string)


def process_sheet(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=None, values_only=False):
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = escape_quotes_in_string(cell.value)


def process_workbook(file_path):
    workbook = load_workbook(file_path)
    for sheet in workbook.worksheets:
        process_sheet(sheet)
    workbook.save(file_path)


def task_func(directory_path='./xlsx_files/'):
    count = 0
    for file_path in glob.glob(os.path.join(directory_path, '*.xlsx')):
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_file_path = os.path.join(temp_dir, os.path.basename(file_path))
                process_workbook(file_path)
                workbook.save(temp_file_path)
                workbook.close()
                os.replace(temp_file_path, file_path)
                count += 1
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")
    return count