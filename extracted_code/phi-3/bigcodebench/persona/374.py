import regex as re
import glob
import os
from openpyxl import load_workbook


def task_func(directory_path='./xlsx_files/'):
    file_count = 0
    for file_name in glob.glob(os.path.join(directory_path, '*.xlsx')):
        file_count += 1
        workbook = load_workbook(file_name)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = re.sub(r'"', r'\"', cell.value)
        workbook.save(file_name)
    return file_count