import os
from pathlib import Path
import pandas as pd
from docx import Document
from openpyxl import load_workbook

def convert_to_csv(file_path, target_dir):
    _, file_extension = os.path.splitext(file_path)
    if file_extension.lower() in ['.txt', '.docx', '.xlsx', '.csv']:
        if file_extension.lower() == '.txt':
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
                df = pd.DataFrame({'Text': [text]})
        elif file_extension.lower() == '.docx':
            doc = Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
            df = pd.DataFrame({'Text': [text]})
        elif file_extension.lower() == '.xlsx':
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            text = sheet.cell(row=1, column=1).value
            df = pd.DataFrame({'Text': [text]})
        elif file_extension.lower() == '.csv':
            df = pd.read_csv(file_path)
        else:
            return

        csv_path = Path(target_dir) / (Path(file_path).stem + '.csv')
        df.to_csv(csv_path, index=False)
        return 1
    return 0


def task_func(source_directory: str, target_directory: str) -> int:
    source_path = Path(source_directory)
    if not source_path.exists():
        raise FileNotFoundError(f"Source directory {source_directory} does not exist.")
    target_path = Path(target_directory)
    if not target_path.exists():
        target_path.mkdir(parents=True)

    files_converted = 0
    for file_path in source_path.glob('*'):
        files_converted += convert_to_csv(str(file_path), target_path