import re
import glob
import os
from openpyxl import load_workbook


def task_func(directory_path='./xlsx_files/'):
    count = 0
    for file_path in glob.glob(os.path.join(directory_path, '*.xlsx')):
        workbook = load_workbook(file_path)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=None, values_only=False):
                for cell in row:
                    if isinstance(cell, str):
                        cell.value = re.sub(r'"', r'\"', cell.value)
        workbook.save(file_path)
        count += 1
    return count